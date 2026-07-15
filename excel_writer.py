from __future__ import annotations

from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

from config import EXCEL_FILE
from database import Database
from logger import logger


class ExcelWriter:
    """
    Escribe las lecturas en la tabla 'GlucosaTabla' de Excel.

    Para cada hora pendiente escribe el último valor disponible
    ANTES o EN esa hora.
    """

    def __init__(self, db: Database):

        self.db = db

        try:
            self.wb = load_workbook(EXCEL_FILE)

        except FileNotFoundError:

            raise FileNotFoundError(
                f"No se encontró el archivo Excel: {EXCEL_FILE}"
            )

        self.sheet = None
        self.table = None

        for ws in self.wb.worksheets:

            if "GlucosaTabla" in ws.tables:

                self.sheet = ws
                self.table = ws.tables["GlucosaTabla"]
                break

        if self.table is None:

            raise RuntimeError(
                "No se encontró la tabla 'GlucosaTabla'."
            )

        (
            self.min_col,
            self.min_row,
            self.max_col,
            self.max_row,
        ) = range_boundaries(self.table.ref)

        logger.info(
            "Tabla localizada (%s)",
            self.table.ref
        )

    ##################################################################

    def sync(self):

        last_sync = self.db.last_excel_sync()

        if last_sync is None:

            all_readings = self.db.get_all()

            if not all_readings:

                logger.info(
                    "No existen lecturas en la base de datos."
                )
                return

            last_sync = (
                all_readings[0]
                .timestamp
                .replace(
                    minute=0,
                    second=0,
                    microsecond=0,
                )
                - timedelta(hours=1)
            )

            logger.info(
                "Primera sincronización."
            )

        current_hour = (
            datetime.now()
            .replace(
                minute=0,
                second=0,
                microsecond=0,
            )
        )

        if current_hour <= last_sync:

            logger.info(
                "Excel ya está sincronizado."
            )
            return

        hours = []

        next_hour = last_sync + timedelta(hours=1)

        while next_hour <= current_hour:

            hours.append(next_hour)

            next_hour += timedelta(hours=1)

        filas_añadidas = 0

        for hour in hours:

            reading = self.db.get_last_before(hour)

            if reading is None:

                logger.warning(
                    "No existe lectura antes de %s",
                    hour,
                )

                continue

            # --------------------------------------------------
            # AQUÍ ESTÁ LA CORRECCIÓN IMPORTANTE
            # --------------------------------------------------

            fecha_str = hour.strftime("%d/%m/%Y")
            hora_str = hour.strftime("%H:%M")

            glucose = reading.glucose

            self.sheet.append(

                [

                    fecha_str,

                    hora_str,

                    glucose,

                ]

            )

            filas_añadidas += 1

            logger.info(

                "Añadida %s %s -> %s mg/dL (lectura real %s)",

                fecha_str,

                hora_str,

                glucose,

                reading.timestamp.strftime("%H:%M:%S"),

            )

        if filas_añadidas == 0:

            logger.info(
                "No había filas nuevas."
            )

            return

        nuevo_final = self.max_row + filas_añadidas

        nuevo_rango = (

            f"{get_column_letter(self.min_col)}{self.min_row}:"

            f"{get_column_letter(self.max_col)}{nuevo_final}"

        )

        self.table.ref = nuevo_rango

        logger.info(
            "Tabla ampliada a %s",
            nuevo_rango,
        )

        self.wb.save(EXCEL_FILE)

        logger.info(
            "Excel guardado."
        )

        self.db.update_last_excel_sync(
            hours[-1]
        )

        logger.info(
            "last_excel_sync actualizado a %s",
            hours[-1],
        )

        logger.info(
            "Sincronización terminada."
        )